import os
import re
from sqlmodel import Session, select
from app.database import engine, create_db_and_tables
from app.models import Unit, Resident, SystemConfig
from openpyxl import load_workbook # Biblioteca para ler Excel

# NOME DO ARQUIVO EXCEL
EXCEL_FILE = "moradores.xlsx"

def clean_name(name_part):
    """Limpa sujeira do nome"""
    if not name_part: return ""
    name_part = str(name_part) # Garante que é string
    # Remove textos entre parenteses
    name_part = re.sub(r'\(.*?\)', '', name_part)
    # Remove palavras chave de ruído
    name_part = name_part.replace("CRIANÇA", "").replace("CRIANÇAS", "").replace("-", "").replace("INQUILINO:", "").replace("INQUILINOS:", "").replace("PROPRIETARIOS:", "")
    # Remove dígitos (idades)
    name_part = re.sub(r'\d+\s*ANOS', '', name_part, flags=re.IGNORECASE)
    name_part = re.sub(r'\d+', '', name_part)
    return name_part.strip()

def import_excel():
    print("🚀 Iniciando importação via EXCEL Direto...")
    
    # 1. Garante banco e config
    create_db_and_tables()
    
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ ERRO: O arquivo '{EXCEL_FILE}' não está na pasta!")
        return

    with Session(engine) as session:
        # Configuração inicial para evitar erro 500
        config = session.exec(select(SystemConfig)).first()
        if not config:
            session.add(SystemConfig(condo_name="Condomínio Bloco 1"))
            session.commit()
            print("⚙️ Configuração inicial criada.")

        # 2. Carrega o Excel
        try:
            wb = load_workbook(EXCEL_FILE, data_only=True) # data_only pega o valor, não a fórmula
            ws = wb.active # Pega a primeira aba
        except Exception as e:
            print(f"❌ Erro ao abrir Excel: {e}")
            return

        # 3. Localiza o Cabeçalho
        header_row_index = None
        col_apto_idx = None
        col_nomes_idx = None

        print("🔎 Procurando colunas APTO e NOMES...")
        
        # Varre as primeiras 20 linhas procurando o cabeçalho
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            row_list = [str(c).upper().strip() if c else '' for c in row]
            
            if "APTO" in row_list and "NOMES" in row_list:
                # O openpyxl usa índice base 1 para linhas, mas aqui estamos iterando
                # Vamos pegar o índice das colunas
                col_apto_idx = row_list.index("APTO")
                col_nomes_idx = row_list.index("NOMES")
                print(f"✅ Cabeçalho encontrado! (Apto: Col {col_apto_idx}, Nomes: Col {col_nomes_idx})")
                break
        
        if col_apto_idx is None:
            print("❌ ERRO: Não encontrei as colunas 'APTO' e 'NOMES'. Verifique a planilha.")
            return

        # 4. Processa os Dados
        count_residents = 0
        
        # Itera sobre as linhas do Excel (começando após o cabeçalho)
        # Como não sabemos a linha exata do cabeçalho pelo iterador acima de forma simples,
        # vamos iterar tudo e usar a lógica de verificar se é dado válido
        
        for row in ws.iter_rows(min_row=1, values_only=True):
            # Pega valores das colunas identificadas
            val_apto = row[col_apto_idx]
            val_nomes = row[col_nomes_idx]

            # Converte para string e limpa
            apto_str = str(val_apto).strip() if val_apto else ""
            nomes_str = str(val_nomes).strip() if val_nomes else ""

            # Pula linhas inválidas ou cabeçalho
            if not apto_str or not nomes_str or apto_str.upper() == "APTO":
                continue
            
            # Tenta converter apto para número/string limpa (ex: 101.0 -> 101)
            try:
                if apto_str.endswith('.0'): apto_str = apto_str[:-2]
            except: pass

            # Cria Unidade
            unit = session.exec(select(Unit).where(Unit.number == apto_str)).first()
            if not unit:
                unit = Unit(block="1", number=apto_str, status="Ocupado")
                session.add(unit)
                session.commit()
                session.refresh(unit)

            # Define Tipo
            profile_type = "PROPRIETARIO"
            if "INQUILINO" in nomes_str.upper():
                profile_type = "INQUILINO"

            # Limpa Nomes
            clean_raw = nomes_str.upper().replace("INQUILINOS:", "").replace("INQUILINO:", "").replace("PROPRIETARIOS:", "")
            names_list = clean_raw.split('/')

            for n in names_list:
                final_name = clean_name(n)
                
                if len(final_name) > 2:
                    # Verifica duplicidade
                    exists = session.exec(select(Resident).where(
                        Resident.unit_id == unit.id,
                        Resident.full_name == final_name.title()
                    )).first()

                    if not exists:
                        res = Resident(
                            unit_id=unit.id,
                            full_name=final_name.title(),
                            profile_type=profile_type,
                            is_active=True,
                            birth_date=None
                        )
                        session.add(res)
                        count_residents += 1
                        print(f"   👤 {apto_str}: {final_name.title()}")

        session.commit()
        print("-" * 30)
        print(f"✅ Importação Concluída via Excel!")
        print(f"👥 Total Moradores: {count_residents}")

if __name__ == "__main__":
    import_excel()